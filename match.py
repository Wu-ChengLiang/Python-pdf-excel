from fuzzywuzzy import process
from openpyxl import load_workbook


#字典数据比较函数
#字典数据比较函数，sql_name是字符串，pdf_name是一个列表，返回一个元组
def find_best_match_threshold(sql_name, pdf_names, threshold):
    best_match = process.extractOne(sql_name, pdf_names)
    if best_match:
        return best_match[0], best_match[1]
    else:
        return None, 0  # 如果没有找到任何匹配项，返回 None 和相似度分数 0

# #两个字符串数据比较函数，两个都是字符串，只返回分数
# def find_best_match_threshold_str(sql_name, pdf_name, threshold):
#     similarity_score = fuzz.ratio(sql_name,pdf_name)
#     if similarity_score >= threshold:
#             return similarity_score
#     else :
#         return 0

# 加载临时映射表
def load_mapping_table(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    mapping_dict = {}
    # 遍历工作表的每一行，从第2行开始（跳过表头）
    for row in sheet.iter_rows(min_row=2, values_only=True):
        original_name = row[4]
        standardized_name = row[5]
        # 检查是否为有效数据
        if original_name is not None and standardized_name is not None:
            mapping_dict[original_name] = standardized_name
        else:
            # 如果某行无效，停止读取后续行
            break
    return mapping_dict


#—————————————————————————————上交所 2.主营业务分产品表格 主营业务分行业、分产品、分地区、分销售模式情况——————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table2_shanghai(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '销售收入'and sql_item['指标单位'] != '%':
            data_category = "营业收入"
        elif sql_item['标准名称'] == '销售收入'and sql_item['指标单位'] == '%':
            data_category = "营业收入比上年增减"
        elif sql_item['标准名称'] == '营业成本'  and sql_item['指标单位'] != '%':
            data_category = "营业成本"
        elif sql_item['标准名称'] == '营业成本'  and sql_item['指标单位'] == '%':
            data_category = "营业成本比上年增减"
        elif sql_item['标准名称'] == '毛利率'  and sql_item['指标单位'] == '%':
            data_category = "毛利率"
        elif sql_item['标准名称'] == '毛利率' and sql_item['指标单位'] == '百分点':
            data_category = "毛利率比上年增减"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table3_shanghai: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table2_shanghai(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目一名称') or item.get('数据类目二名称') or item.get('数据类目三名称') or '默认值'): item
        for item in sql_filtered
    }

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['数据类目名称']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表

    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())

    for pdf_name in pdf_names:
        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table2_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            if get_high is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table2_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '高匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            if get_medium is True:
                unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
                unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table2_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '中匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            if get_low is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table2_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '低匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names

#——————————————————————————————————————————————3.产销表 ————————————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table3_shanghai(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '产量'and sql_item['指标单位'] != '%':
            data_category = "生产量"
        elif sql_item['标准名称'] == '销量'  and sql_item['指标单位'] != '%':
            data_category = "销售量"
        elif sql_item['标准名称'] == '库存量'and sql_item['指标单位'] != '%':
            data_category = "库存量"
        elif sql_item['标准名称'] == '产量'  and sql_item['指标单位'] == '%':
            data_category = "生产量比上年增减"
        elif sql_item['标准名称'] == '销量'  and sql_item['指标单位'] == '%':
            data_category = "销售量比上年增减"
        elif sql_item['标准名称'] == '库存量' and sql_item['指标单位'] == '%':
            data_category = "库存量比上年增减"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table3_shanghai: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table3_shanghai(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目一名称') or item.get('数据类目二名称') or item.get('数据类目三名称') or '默认值'): item
        for item in sql_filtered
    }
    sql_names = list(sql_map.keys())

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['主要产品']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称

    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"

    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表


    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())

    for pdf_name in pdf_names:
        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table3_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table3_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '高匹配度'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
            if get_high is True:
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
            unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table3_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '中匹配度'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
            if get_medium is True:
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table3_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '低匹配度'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
            if get_low is True:
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")
    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names




#——————————————————————————————————————————————4.成本分析表 ————————————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table4_shanghai(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '成本' and sql_item['指标单位'] != '%':
            data_category = "本期金额"
        elif sql_item['标准名称'] == '成本占比'  and sql_item['指标单位'] == '%':
            data_category = "本期占总成本比例"
        elif sql_item['标准名称'] == '成本'and sql_item['指标单位'] == '%':
            data_category = "本期金额较上年同期变动比例"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table3_shanghai: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }


def match_table4_shanghai(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10,threshold_double=80, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    sql_map = {
        (
            (item.get('数据类目一名称') or '默认值'),
            ('小计' if item.get('数据类目三名称') == None else item.get('数据类目三名称') or '小计')
            # (item.get('数据类目三名称') or 'None')
        ): item
        for item in sql_filtered
    }
    sql_names = list(sql_map.keys())  # 获取所有 SQL 名称


    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {
        ((item.get('分产品') or '默认值'),(item.get('成本构成项目') or '默认值')): item
        for item in pdf_dic
    }
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    def fuzzy_match(str1, str2):
        from difflib import SequenceMatcher
        # 计算相似度比率，然后乘以 100 转换为百分比形式
        score = SequenceMatcher(None, str1, str2).ratio() * 100
        # 四舍五入到三位小数
        return round(score, 1)

    def match_tuple(pdf_name, sql_names, threshold=80):
        best_match = None
        best_score = 0
        # 特殊情况：PDF 数据的第二个字段是“合计”
        if pdf_name[1] == '合计':
            for sql_name in sql_names:
                if sql_name[0] == '默认值' and sql_name[1] == '小计':  # 匹配 SQL 中的“合计”
                    return sql_name, 100  # 直接返回匹配结果，相似度为100
        # 特殊情况：PDF 数据的第一个字段是“合计” 大合计
        if pdf_name[0] == '合计':
            for sql_name in sql_names:
                if sql_name[0] == "产品合计" and sql_name[1] == '小计':
                    return sql_name, 100

        # if pdf_name[1] == '小计':
        #     for sql_name in sql_names:
        #         if sql_name[1] == None:  # 匹配 SQL 中的“None”
        #             score = fuzzy_match(pdf_name[0], sql_name[0])
        #             if score >= threshold:
        #                 return sql_name, score

        # 正常匹配逻辑
        else:
            for sql_name in sql_names:
                score1 = fuzzy_match(pdf_name[0], sql_name[0])
                score2 = fuzzy_match(pdf_name[1], sql_name[1])

                # 如果两个字段的相似度都低于阈值，直接跳过
                if (score1 < threshold_double ) and (score2 < threshold_double ):
                    continue

                # 计算综合相似度
                combined_score = (score1 + score2) / 2

                # 如果综合相似度高于当前的最佳分数，更新最佳匹配
                if combined_score > best_score:
                    best_score = combined_score
                    best_match = sql_name
        return best_match, round(best_score, 1)


    for pdf_name in pdf_names:
        match_name, score = match_tuple(pdf_name, sql_names, threshold=threshold)

        # 先排除空值，如果 SQL 未匹配到任何结果
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)  # 记录未匹配的 PDF 名称
            unmatched_sql_names.add(match_name)  # 记录未匹配的 SQL 名称
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 最佳匹配相似度: {score}")
            continue

        # 完全匹配（相似度 > 99）
        if score > 99:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table4_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if not message_only_wrong:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 相似度: {score}")

        # 高匹配度（threshold <= score <= 99）
        elif score >= threshold and score <= 99:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table4_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '高匹配度'
            matched_data['pdf原始名称'] = pdf_name

            if get_high:  # 根据参数决定是否记录高匹配度结果
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if not message_only_wrong:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 相似度: {score}")

        # 中匹配度（low_threshold <= score < threshold）
        elif score >= low_threshold and score < threshold:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table4_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '中匹配度'
            matched_data['pdf原始名称'] = pdf_name

            if get_medium:  # 根据参数决定是否记录中匹配度结果
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 最佳匹配相似度: {score}")

        # 低匹配度（score < low_threshold）
        else:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table4_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '低匹配度'
            matched_data['pdf原始名称'] = pdf_name

            if get_low:  # 根据参数决定是否记录低匹配度结果
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 最佳匹配相似度: {score}")

        # 输出未匹配的 SQL 名称
    if unmatched_sql_names:
            print(f"❗成本分析表 未匹配的 SQL 条目: {RED}{unmatched_sql_names}{RESET}")
    else:
            print("成本分析表 ❗ 所有 SQL 条目均已匹配  ❗ 或不存在")

        # 输出未匹配的 PDF 名称
    if unmatched_pdf_names:
            print(f"❗成本分析表 未匹配的 PDF 条目: {RED}{unmatched_pdf_names}{RESET}")
    else:
            print("成本分析表 ❗ 所有 SQL 条目均已匹配  ❗ 或不存在")

    return mix_dic, unmatched_pdf_names, unmatched_sql_names



#—————————————————————————————————————————————5.前五大客户表 公司前五名客户表————————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table5_shanghai(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '销售收入'and sql_item['指标单位'] != '%':
            data_category = "销售额"
        elif sql_item['标准名称'] == '销售收入占比'  and sql_item['指标单位'] == '%':
            data_category = "占年度销售总额比例"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table5_shanghai: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table5_shanghai(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目三名称') or '默认值'): item #item.get('数据类目一名称') or item.get('数据类目二名称') or
        for item in sql_filtered
    }
    sql_names = list(sql_map.keys())

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['客户名称']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表

    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())

    for pdf_name in pdf_names:
        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table5_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table5_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '高匹配度'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
            if get_high is True:
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
            unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

            if get_medium is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table5_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '中匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            if get_low is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table5_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '低匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")
    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names





#——————————————————————————————————————————6.前五大供应商表 公司前五名供应商表————————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table6_shanghai(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '采购金额' and sql_item['指标单位'] != '%':
            data_category = "采购额"
        elif sql_item['标准名称'] == '采购金额占比'  and sql_item['指标单位'] == '%':
            data_category = "占年度采购总额比例"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table6_shanghai: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table6_shanghai(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目三名称') or '默认值'): item #item.get('数据类目一名称') or item.get('数据类目二名称') or
        for item in sql_filtered
    }
    sql_names = list(sql_map.keys())

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['供应商名称']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表

    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())

    for pdf_name in pdf_names:
        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table6_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table6_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '高匹配度'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
            if get_high is True:
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
            unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

            if get_medium is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table6_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '中匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            if get_low is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table6_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '低匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")
    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names

#——————————————————————————————————————————————7.专利表  7.报告期内获得的知识产权列表 上海————————————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table7_shanghai(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '新增知识产权申请数量':
            data_category = "本年新增申请数"
        elif sql_item['标准名称'] == '新增知识产权有效数量':
            data_category = "本年新增获得数"
        elif sql_item['标准名称'] == '累计知识产权申请数量':
            data_category = "累计数量申请数"
        elif sql_item['标准名称'] == '累计知识产权有效数量':
            data_category = "累计数量获得数"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table2_shanghai: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table7_shanghai(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目一名称') or '默认值'): item
        for item in sql_filtered
    }

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['专利类型']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表

    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())

    for pdf_name in pdf_names:
        #这里要打一个补丁：也就是当数据类目一 “默认值”= “产品合计”

        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table7_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            if get_high is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table7_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '高匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            if get_medium is True:
                unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
                unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table7_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '中匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            if get_low is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table7_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '低匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names





#—————————————————————————————深交所 2.主营业务分产品表格 主营业务分行业、分产品、分地区、分销售模式情况——————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table2_shenzhen(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '销售收入'and sql_item['指标单位'] != '%':
            data_category = "营业收入"
        elif sql_item['标准名称'] == '销售收入'and sql_item['指标单位'] == '%':
            data_category = "营业收入比上年增减"

        else:
            print(f"sql_item没有正确匹配catagory_choose_table3_shenzhen: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table2_shenzhen(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目一名称') or item.get('数据类目二名称') or item.get('数据类目三名称') or '默认值'): item
        for item in sql_filtered
    }

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['数据类目名称']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表

    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name
    # print(temp_sql_map)

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())
    # print(temp_sql_names)

    for pdf_name in pdf_names:
        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table2_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            if get_high is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table2_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '高匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            if get_medium is True:
                unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
                unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table2_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '中匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            if get_low is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table2_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '低匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names

#—————————————————————————————深交所 2.5 主营业务分产品表格 主营业务分行业、分产品、分地区、分销售模式情况——————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table2_5_shenzhen(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '销售收入' and sql_item['指标单位'] != '%':
            data_category = "营业收入"
        elif sql_item['标准名称'] == '销售收入' and sql_item['指标单位'] == '%':
            data_category = "营业收入比上年增减"
        elif sql_item['标准名称'] == '营业成本' and sql_item['指标单位'] != '%':
            data_category = "营业成本"
        elif sql_item['标准名称'] == '营业成本' and sql_item['指标单位'] == '%':
            data_category = "营业成本比上年增减"
        elif sql_item['标准名称'] == '毛利率' and sql_item['指标单位'] == '%':
            data_category = "毛利率"
        elif sql_item['标准名称'] == '毛利率' and sql_item['指标单位'] == '百分点':
            data_category = "毛利率比上年增减"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table3_shenzhen: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table2_5_shenzhen(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目一名称') or item.get('数据类目二名称') or item.get('数据类目三名称') or '默认值'): item
        for item in sql_filtered
    }

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['数据类目名称']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表

    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name
    # print(temp_sql_map)

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())
    # print(temp_sql_names)

    for pdf_name in pdf_names:
        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table2_5_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            if get_high is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table2_5_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '高匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            if get_medium is True:
                unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
                unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table2_5_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '中匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            if get_low is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table2_5_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '低匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names


#——————————————————————————————————————————————3.产销表 ————————————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table3_shenzhen(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '产量' and sql_item['指标单位'] != '%':
            data_category = "生产量、销售量、库存量"
        elif sql_item['标准名称'] == '销量'  and sql_item['指标单位'] != '%':
            data_category = "生产量、销售量、库存量"
        elif sql_item['标准名称'] == '库存量'and sql_item['指标单位'] != '%':
            data_category = "生产量、销售量、库存量"
        elif sql_item['标准名称'] == '产量'  and sql_item['指标单位'] == '%':
            data_category = "生产量、销售量、库存量比上年增减"
        elif sql_item['标准名称'] == '销量'  and sql_item['指标单位'] == '%':
            data_category = "生产量、销售量、库存量比上年增减"
        elif sql_item['标准名称'] == '库存量' and sql_item['指标单位'] == '%':
            data_category = "生产量、销售量、库存量比上年增减"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table3_shenzhen: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

        # 匹配的时候没有正常的进入数据，1.注释掉，写一个最简单不出错的的匹配算法，看看结果，然后增加鲁棒性

def match_table3_shenzhen(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10,threshold_double=80, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    sql_map = {
        (
            (item.get('数据类目一名称') or '默认值'),
            (item.get('标准名称') or '默认值'),
        ): item
        for item in sql_filtered
    }
    sql_names = list(sql_map.keys())  # 获取所有 SQL 名称


    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {
        ((item.get('行业分类') or '默认值'),(item.get('项目') or '默认值')): item
        for item in pdf_dic
    }
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称




    def fuzzy_match(str1, str2):
        #计算两个字符串的匹配度
        from difflib import SequenceMatcher
        # 计算相似度比率，然后乘以 100 转换为百分比形式
        score = SequenceMatcher(None, str1, str2).ratio() * 100
        # 四舍五入到三位小数
        return round(score, 1)

    def match_tuple(pdf_name, sql_names, threshold=80):
        best_match = None
        best_score = 0
        # 特殊情况：PDF 数据的第一个字段是“合计”，等待实例测试
        if pdf_name[0] == '合计':
            for sql_name in sql_names:
                if sql_name[1] == '默认值':  # 匹配 SQL 中的“合计”
                    return sql_name, 100  # 直接返回匹配结果，相似度为100

        # 正常匹配逻辑
        else:
            # 先用sql_names 的第二个元组和 pdf_names 的第二个元组相匹配

                #实际的核心匹配步骤
                #下一排其实已经按照pdf去循环了
                for sql_name in sql_names:
                    score1 = fuzzy_match(pdf_name[0], sql_name[0])
                    score2 = fuzzy_match(pdf_name[1], sql_name[1])
                    # 如果标准名称的相似度都低于阈值，直接跳过
                    if  score2 < threshold_double :
                        continue

                    if score1 > threshold and score2 >= threshold:
                        best_score = score1
                        best_match = sql_name
        return best_match, round(best_score, 1)




    for pdf_name in pdf_names:
        match_name, score = match_tuple(pdf_name, sql_names, threshold=threshold)

        # 先排除空值，如果 SQL 未匹配到任何结果
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)  # 记录未匹配的 PDF 名称
            unmatched_sql_names.add(match_name)  # 记录未匹配的 SQL 名称

            print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 最佳匹配相似度: {score}")
            continue

        # 完全匹配（相似度 > 99）
        if score > 99:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table3_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if not message_only_wrong:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 相似度: {score}")

        # 高匹配度（threshold <= score <= 99）
        elif score >= threshold and score <= 99:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table3_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '高匹配度'
            matched_data['pdf原始名称'] = pdf_name

            if get_high:  # 根据参数决定是否记录高匹配度结果
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if not message_only_wrong:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 相似度: {score}")

        # 中匹配度（low_threshold <= score < threshold）
        elif score >= low_threshold and score < threshold:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table3_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '中匹配度'
            matched_data['pdf原始名称'] = pdf_name

            if get_medium:  # 根据参数决定是否记录中匹配度结果
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 最佳匹配相似度: {score}")

        # 低匹配度（score < low_threshold）
        else:
            unmatched_pdf_names.add(pdf_name)  # 记录未匹配的 PDF 名称
            unmatched_sql_names.add(match_name)  # 记录未匹配的 SQL 名称

            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table3_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '低匹配度'
            matched_data['pdf原始名称'] = pdf_name

            if get_low:  # 根据参数决定是否记录低匹配度结果
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 最佳匹配相似度: {score}")

        # 输出未匹配的 SQL 名称
    if unmatched_sql_names:
            print(f"❗产销表 未匹配的 SQL 条目: {RED}{unmatched_sql_names}{RESET}")
    else:
            print("产销表 ❗ 所有 SQL 条目均已匹配  ❗ 或不存在")

        # 输出未匹配的 PDF 名称
    if unmatched_pdf_names:
            print(f"❗产销表 未匹配的 PDF 条目: {RED}{unmatched_pdf_names}{RESET}")
    else:
            print("产销表 ❗ 所有 SQL 条目均已匹配  ❗ 或不存在")

    return mix_dic, unmatched_pdf_names, unmatched_sql_names


# #核心匹配函数,这个是上交所版本的，深交所需要改变逻辑
# def match_table3_shenzhen(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
#     mix_dic = []  # 用于存储匹配成功的结果
#     unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
#     unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称
#
#     # 数据结构方案
#     sql_map = {
#         (item.get('数据类目一名称') or item.get('数据类目二名称') or item.get('数据类目三名称') or '默认值'): item
#         for item in sql_filtered
#     }
#     sql_names = list(sql_map.keys())
#
#     # 提取 PDF 数据中的名称，并创建映射
#     pdf_map = {item['主要产品']: item for item in pdf_dic}
#     pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称
#
#     # 定义颜色输出
#     GREEN = "\033[92m"
#     RED = "\033[91m"
#     RESET = "\033[0m"
#
#     if mapping_file_path:
#         mapping_table = load_mapping_table(mapping_file_path)
#     else:
#         mapping_table = {}  # 如果未提供文件路径，则使用空映射表
#
#
#     # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
#     temp_sql_map = {}
#     original_name_map = {}  # 用于记录临时名称对应的原始名称
#     for original_name, item in sql_map.items():
#         temp_name = original_name
#         for old, new in mapping_table.items():
#             temp_name = temp_name.replace(old, new)
#         temp_sql_map[temp_name] = item
#         original_name_map[temp_name] = original_name
#
#     # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
#     temp_sql_names = list(temp_sql_map.keys())
#
#     for pdf_name in pdf_names:
#         # 查找最佳匹配的 SQL 名称
#         match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)
#
#         # 先排除空值，这个情况往往是EP输入错误，sql无数据
#         if match_name is None:
#             unmatched_pdf_names.add(pdf_name)
#             unmatched_sql_names.add(match_name)
#             print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
#             continue
#
#         # 获取原始名称
#         # original_match_name = original_name_map[match_name]
#         original_match_name = original_name_map.get(match_name)
#
#         # 相同
#         if score > 99:  # 100
#             matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
#             matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
#             matched_data = catagory_choose_table3_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果
#
#             matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
#             matched_data['精度'] = '相同'
#             matched_data['pdf原始名称'] = pdf_name
#             matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
#
#             mix_dic.append(matched_data)  # 将匹配结果添加到列表中
#             if message_only_wrong is False:
#                 print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")
#
#         # 高精度
#         if score >= threshold and score <= 99:  # 80~99
#             matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
#             matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
#             matched_data = catagory_choose_table3_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果
#
#             matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
#             matched_data['精度'] = '高匹配度'
#             matched_data['pdf原始名称'] = pdf_name
#             matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
#             if get_high is True:
#                 mix_dic.append(matched_data)  # 将匹配结果添加到列表中
#                 if message_only_wrong is False:
#                     print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")
#
#         # 中精度
#         if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
#             unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
#             unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称
#
#             matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
#             matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
#             matched_data = catagory_choose_table3_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果
#
#             matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
#             matched_data['精度'] = '中匹配度'
#             matched_data['pdf原始名称'] = pdf_name
#             matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
#             if get_medium is True:
#                 mix_dic.append(matched_data)  # 将匹配结果添加到列表中
#                 print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")
#
#         # 低精度
#         if score < low_threshold:
#             matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
#             matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
#             matched_data = catagory_choose_table3_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果
#
#             matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
#             matched_data['精度'] = '低匹配度'
#             matched_data['pdf原始名称'] = pdf_name
#             matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
#             if get_low is True:
#                 mix_dic.append(matched_data)  # 将匹配结果添加到列表中
#                 print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")
#     print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
#     return mix_dic, unmatched_pdf_names, unmatched_sql_names




#——————————————————————————————————————————————4.成本分析表 ————————————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table4_shenzhen(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '成本' and sql_item['指标单位'] != '%':
            data_category = "本期金额"
        elif sql_item['标准名称'] == '成本占比'  and sql_item['指标单位'] == '%':
            data_category = "本期占总成本比例"
        elif sql_item['标准名称'] == '成本'and sql_item['指标单位'] == '%':
            data_category = "本期金额较上年同期变动比例"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table3_shenzhen: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }


def match_table4_shenzhen(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10,threshold_double=80, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    sql_map = {
        (
            (item.get('数据类目一名称') or '默认值'),
            # ('小计' if item.get('数据类目三名称') == None else item.get('数据类目三名称') or '小计')
            (item.get('数据类目三名称') or '小计')
        ): item
        for item in sql_filtered
    }
    sql_names = list(sql_map.keys())  # 获取所有 SQL 名称


    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {
        ((item.get('分产品') or '默认值'),(item.get('成本构成项目') or '默认值')): item
        for item in pdf_dic
    }
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    def fuzzy_match(str1, str2):
        from difflib import SequenceMatcher
        # 计算相似度比率，然后乘以 100 转换为百分比形式
        score = SequenceMatcher(None, str1, str2).ratio() * 100
        # 四舍五入到三位小数
        return round(score, 1)

    def match_tuple(pdf_name, sql_names, threshold=80):
        best_match = None
        best_score = 0
        # 特殊情况：PDF 数据的第二个字段是“合计”
        if pdf_name[1] == '合计':
            for sql_name in sql_names:
                if sql_name[0] == '默认值' and sql_name[1] == '小计':  # 匹配 SQL 中的“合计”
                    return sql_name, 100  # 直接返回匹配结果，相似度为100

        # if pdf_name[1] == '小计':
        #     for sql_name in sql_names:
        #         if sql_name[1] == None:  # 匹配 SQL 中的“None”
        #             score = fuzzy_match(pdf_name[0], sql_name[0])
        #             if score >= threshold:
        #                 return sql_name, score

        # 正常匹配逻辑
        else:
            for sql_name in sql_names:
                score1 = fuzzy_match(pdf_name[0], sql_name[0])
                score2 = fuzzy_match(pdf_name[1], sql_name[1])

                # 如果两个字段的相似度都低于阈值，直接跳过
                if (score1 < threshold_double ) and (score2 < threshold_double ):
                    continue

                # 计算综合相似度
                combined_score = (score1 + score2) / 2

                # 如果综合相似度高于当前的最佳分数，更新最佳匹配
                if combined_score > best_score:
                    best_score = combined_score
                    best_match = sql_name
        return best_match, round(best_score, 1)


    for pdf_name in pdf_names:
        match_name, score = match_tuple(pdf_name, sql_names, threshold=threshold)

        # 先排除空值，如果 SQL 未匹配到任何结果
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)  # 记录未匹配的 PDF 名称
            unmatched_sql_names.add(match_name)  # 记录未匹配的 SQL 名称
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 最佳匹配相似度: {score}")
            continue

        # 完全匹配（相似度 > 99）
        if score > 99:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table4_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if not message_only_wrong:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 相似度: {score}")

        # 高匹配度（threshold <= score <= 99）
        elif score >= threshold and score <= 99:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table4_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '高匹配度'
            matched_data['pdf原始名称'] = pdf_name

            if get_high:  # 根据参数决定是否记录高匹配度结果
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if not message_only_wrong:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 相似度: {score}")

        # 中匹配度（low_threshold <= score < threshold）
        elif score >= low_threshold and score < threshold:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table4_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '中匹配度'
            matched_data['pdf原始名称'] = pdf_name

            if get_medium:  # 根据参数决定是否记录中匹配度结果
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 最佳匹配相似度: {score}")

        # 低匹配度（score < low_threshold）
        else:
            matched_sql = sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table4_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '低匹配度'
            matched_data['pdf原始名称'] = pdf_name

            if get_low:  # 根据参数决定是否记录低匹配度结果
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET}, 最佳匹配相似度: {score}")

        # 输出未匹配的 SQL 名称
    if unmatched_sql_names:
            print(f"❗成本分析表 未匹配的 SQL 条目: {RED}{unmatched_sql_names}{RESET}")
    else:
            print("成本分析表 ❗ 所有 SQL 条目均已匹配  ❗ 或不存在")

        # 输出未匹配的 PDF 名称
    if unmatched_pdf_names:
            print(f"❗成本分析表 未匹配的 PDF 条目: {RED}{unmatched_pdf_names}{RESET}")
    else:
            print("成本分析表 ❗ 所有 SQL 条目均已匹配  ❗ 或不存在")

    return mix_dic, unmatched_pdf_names, unmatched_sql_names



#—————————————————————————————————————————————5.前五大客户表 公司前五名客户表————————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table5_shenzhen(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '销售收入'and sql_item['指标单位'] != '%':
            data_category = "销售额"
        elif sql_item['标准名称'] == '销售收入占比'  and sql_item['指标单位'] == '%':
            data_category = "占年度销售总额比例"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table5_shenzhen: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table5_shenzhen(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目三名称') or '默认值'): item #item.get('数据类目一名称') or item.get('数据类目二名称') or
        for item in sql_filtered
    }
    sql_names = list(sql_map.keys())

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['客户名称']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表

    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())

    for pdf_name in pdf_names:
        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table5_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table5_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '高匹配度'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
            if get_high is True:
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
            unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

            if get_medium is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table5_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '中匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            if get_low is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table5_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '低匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")
    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names





#——————————————————————————————————————————6.前五大供应商表 公司前五名供应商表————————————————————————————————————————————————————————
#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table6_shenzhen(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '采购金额' and sql_item['指标单位'] != '%':
            data_category = "采购额"
        elif sql_item['标准名称'] == '采购金额占比'  and sql_item['指标单位'] == '%':
            data_category = "占年度采购总额比例"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table6_shenzhen: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table6_shenzhen(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目三名称') or '默认值'): item #item.get('数据类目一名称') or item.get('数据类目二名称') or
        for item in sql_filtered
    }
    sql_names = list(sql_map.keys())

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['供应商名称']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表

    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())

    for pdf_name in pdf_names:
        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table6_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table6_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '高匹配度'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
            if get_high is True:
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
            unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

            if get_medium is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table6_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '中匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            if get_low is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table6_shenzhen(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '低匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")
    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names



#——————————————————————————————————————————————7.专利表  7.报告期内获得的知识产权列表 上海————————————————————————————————————————————————————————————

#标准名称的动态选择，筛选出对应的pdf字段；使用字典解包操作**sql_item，将SQL数据全部包含到结果字典中；最后把其他需要赋值的日期加入到数据中
def catagory_choose_table7_shanghai(sql_item, pdf_data, JZRQ, XXFBRQ, XXLL):
        if sql_item['标准名称'] == '新增知识产权申请数量':
            data_category = "新增知识产权申请数量"
        elif sql_item['标准名称'] == '新增知识产权有效数量':
            data_category = "新增知识产权有效数量"
        elif sql_item['标准名称'] == '累计知识产权申请数量':
            data_category = "累计知识产权申请数量"
        elif sql_item['标准名称'] == '累计知识产权有效数量':
            data_category = "累计知识产权有效数量"
        else:
            print(f"sql_item没有正确匹配catagory_choose_table2_shanghai: {sql_item}")
            return {}
        return {
            **sql_item,
            '截止日期': JZRQ,
            '信息发布日期': XXFBRQ,
            '信息来源': XXLL,
            '指标数据': pdf_data.get(data_category)
        }

#核心匹配函数
def match_table7_shanghai(sql_filtered, pdf_dic, JZRQ, XXFBRQ, XXLL, threshold=80, low_threshold=10, get_high=True, get_medium=True, get_low=False, message_only_wrong=False,mapping_file_path=None):
    # 定义颜色输出
    GREEN = "\033[92m"
    RED = "\033[91m"
    RESET = "\033[0m"
    mix_dic = []  # 用于存储匹配成功的结果
    unmatched_pdf_names = set()  # 用于记录未匹配的 PDF 名称
    unmatched_sql_names = set()  # 用于记录未匹配的 SQL 名称

    # 数据结构方案
    sql_map = {
        (item.get('数据类目一名称') or '默认值'): item
        for item in sql_filtered
    }

    # 提取 PDF 数据中的名称，并创建映射
    pdf_map = {item['专利类型']: item for item in pdf_dic}
    pdf_names = list(pdf_map.keys())  # 获取所有 PDF 名称


    if mapping_file_path:
        mapping_table = load_mapping_table(mapping_file_path)
    else:
        mapping_table = {}  # 如果未提供文件路径，则使用空映射表

    # 创建临时映射表，temp存储new+原始数据，original存储new+原始名称的值
    temp_sql_map = {}
    original_name_map = {}  # 用于记录临时名称对应的原始名称
    for original_name, item in sql_map.items():
        temp_name = original_name
        for old, new in mapping_table.items():
            temp_name = temp_name.replace(old, new)
        temp_sql_map[temp_name] = item
        original_name_map[temp_name] = original_name

    # 使用临时映射表的键作为匹配的目标，temp_sql_names就是现在拿来做匹配的名称
    temp_sql_names = list(temp_sql_map.keys())

    for pdf_name in pdf_names:
        #这里要打一个补丁：也就是当数据类目一 “默认值”= “产品合计”

        # 查找最佳匹配的 SQL 名称
        match_name, score = find_best_match_threshold(pdf_name, temp_sql_names, threshold)

        # 先排除空值，这个情况往往是EP输入错误，sql无数据
        if match_name is None:
            unmatched_pdf_names.add(pdf_name)
            unmatched_sql_names.add(match_name)
            print(f"{RED}sql无数据，导致未匹配成功: PDF {pdf_name} -> SQL {match_name}{RESET},最佳匹配相似度: {score}")
            continue

        # 获取原始名称
        # original_match_name = original_name_map[match_name]
        original_match_name = original_name_map.get(match_name)

        # 相同
        if score > 99:  # 100
            matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
            matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
            matched_data = catagory_choose_table7_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

            matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
            matched_data['精度'] = '相同'
            matched_data['pdf原始名称'] = pdf_name
            matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

            mix_dic.append(matched_data)  # 将匹配结果添加到列表中
            if message_only_wrong is False:
                print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 高精度
        if score >= threshold and score <= 99:  # 80~99
            if get_high is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table7_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '高匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中

                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                if message_only_wrong is False:
                    print(f"{GREEN}匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET}, 相似度: {score}")

        # 中精度
        if score < threshold and score >= low_threshold:  # 也就是分数大于low_threshold(10)且小于80
            if get_medium is True:
                unmatched_pdf_names.add(pdf_name)  # 从未匹配的 PDF 名称中移除已匹配的名称
                unmatched_sql_names.add(original_match_name)  # 从未匹配的 SQL 名称中移除已匹配的名称

                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table7_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '中匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

        # 低精度
        if score < low_threshold:
            if get_low is True:
                matched_sql = temp_sql_map[match_name]  # 获取匹配的 SQL 数据
                matched_pdf = pdf_map[pdf_name]  # 获取当前 PDF 数据
                matched_data = catagory_choose_table7_shanghai(matched_sql, matched_pdf, JZRQ, XXFBRQ, XXLL)  # 处理匹配结果

                matched_data['相似度分数'] = score  # 添加相似度分数到匹配结果中
                matched_data['精度'] = '低匹配度'
                matched_data['pdf原始名称'] = pdf_name
                matched_data['sql原始名称'] = original_match_name  # 添加原始名称到结果中
                mix_dic.append(matched_data)  # 将匹配结果添加到列表中
                print(f"{RED}未匹配成功: PDF {pdf_name} -> SQL {original_match_name}{RESET},最佳匹配相似度: {score}")

    print("⚡️⚡️️⚡️⚡️⚡️⚡️⚡️  (๑•̀ㅂ•́)ง✧  ⚡️⚡️️⚡️⚡️⚡️⚡️⚡️")
    return mix_dic, unmatched_pdf_names, unmatched_sql_names






