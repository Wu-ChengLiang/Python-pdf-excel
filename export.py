import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font


def append_to_excel(data, output_file, columns, message="数据传输完成"):
    # 检查文件是否存在
    if os.path.exists(output_file):
        # 读取现有文件
        existing_df = pd.read_excel(output_file, engine='openpyxl')
        # 获取最大行号
        start_row = len(existing_df)
    else:
        # 如果文件不存在，从第一行开始写入
        start_row = 0
        existing_df = pd.DataFrame(columns=columns)  # 创建一个空的DataFrame，包含表头字段


    # 将新数据转换为DataFrame
    new_df = pd.DataFrame(data, columns=columns)

    # # 删除 existing_df 和 new_df 中的全空值列
    # existing_df = existing_df.dropna(axis=1, how='all')  # 删除全为空值的列
    # new_df = new_df.dropna(axis=1, how='all')  # 删除全为空值的列

    # 将新数据追加到现有数据中
    updated_df = pd.concat([existing_df, new_df], ignore_index=True)

    # 将更新后的数据保存到Excel文件
    updated_df.to_excel(output_file, index=False, engine='openpyxl')


    # 打印消息   检查 message 是否有内容
    if message.strip():  # 使用 strip() 去除可能的空白字符，确保消息非空
        print(f"{message}🗺️ {output_file} 🗺️🌍🌍🌍🌍🌍")
    else:
        pass # 可选：打印提示信息，或者直接不打印任何内容


def sql_export(output_file,data,JJRQ,EPBH,XXLL_processed):
    # 标准格式，不需要颜色注记和附加的三列，但是保留报错提示
    columns = [
        "EP编号", "行业代码", "信息来源", "信息来源编码", "信息发布日期", "截止日期", "财政年度",
        "经营业务类型代码", "数据类目一", "数据类目一名称", "数据类目一代码", "数据类目二",
        "数据类目二名称", "数据类目二代码", "数据类目三", "数据类目三名称", "主体原始名称",
        "指标代码", "标准名称", "指标名称", "指标数据", "指标单位", "匹配代码-单位",
        "统计口径", "统计期间", "页码", "指标内容", "是否有效", "备注说明", "行编码"
    ]

    #数据的列顺序和名称与预期一致
    data = pd.DataFrame(data,columns=columns)

    #数据筛选：按日期   最新的日期(JJRQ)参数的  = sql["截止日期"]
    data_filtered = data[data["截止日期"] == f"{JJRQ}"]

    #统计sql真正的页码个数
    max_page = len(data_filtered)
    #按照页码排序
    data_sort = data_filtered.sort_values(by=["页码"], ascending=True)

    #excel导出
    #在 Python 中，路径分隔符可以使用 os.path.join 或 pathlib 来处理，以避免跨平台问题。
    #output_file_path = os.path.join(output_file, f"SQL_{EPBH}{XXLL_processed}.xlsx")

    output_file_path = f"{output_file}\\SQL_{EPBH}{XXLL_processed}.xlsx"

    if data_sort.empty:
        print("SQL数据为空")
    else:
        data_sort.to_excel(output_file_path, index=False, engine='openpyxl')
        print(f"1.SQL数据已保存到: {output_file_path}")

    return max_page

def highlight_and_clean_excel(max_page,input_file):
    """
    对Excel文件进行数据清洗和高亮处理，并保存到原始路径。
    """
    # 数据清洗
    try:
        # 读取原始数据
        df_before = pd.read_excel(input_file, engine='openpyxl')
    except Exception as e:
        print(f"读取文件时出错：{e}")
        return

    # # # 按“相似度分数”和优先级（中文和0的优先级更加低降序排列，并保留每个“页码”的第一条记录
    #数据筛选更新
    # 定义函数：判断指标数据是否为0或无法转换为数字
    def assign_priority(value):
        try:
            num = float(value)  # 尝试将值转换为数字
            if num == 0:  # 如果值为0，优先级低
                return 1
            else:
                return 0  # 数字且不为0，优先级高
        except ValueError:  # 如果无法转换为数字（如中文），优先级低
            return 1

    # 添加辅助列：priority
    df_before["priority"] = df_before["指标数据"].apply(assign_priority)

    # 按优先级和相似度分数排序
    # 先按优先级升序（低优先级排在后面），再按相似度分数降序
    df_sorted = df_before.sort_values(by=["priority", "相似度分数"], ascending=[True, False])

    #去重有一个很关键的bug，因为是根据页码去重，所以如果页码都是一致的，会只保留一条数据
    #但是如果去掉去重功能，会让数据冗余量大增；如何能够很精准的识别页码是不是都一致呢？
    #其实只要pd读取页码的数据，然后如果同一个页码超过5条，或是页码全都是空值，那么很简单：if else 就可以了

    # 检查页码是否全为空值
    all_pages_empty = df_sorted["页码"].isna().all()
    # 检查是否有页码重复超过7条
    page_counts = df_sorted["页码"].value_counts()
    any_page_over_7 = (page_counts > 7 ).any()

    # 根据条件判断是否去重
    if all_pages_empty or any_page_over_7:
        # 如果页码全为空或有页码重复超过5条，保留去重逻辑
        df_s = df_sorted.copy()
    else :
        # 如果不满足条件，保留所有数据
        df_s = df_sorted.drop_duplicates(subset="页码", keep = "first")

    # # 去重：保留每个“页码”的第一条记录
    # df_s = df_sorted.drop_duplicates(subset="页码", keep="first")

    # 删除辅助列
    df = df_s.drop(columns=["priority"])

    # 统计被清除的记录数
    out = len(df_sorted) - len(df_s)

    # 计算冗余率
    all_records = len(df_before)
    redundancy_rate = out / all_records if all_records > 0 else 0
    print(f"冗余率：{redundancy_rate:.2%}")

    # 数据分析
    get = df['页码'].nunique()
    get_high = df[df['相似度分数'] > 70]['页码'].nunique()

    # 计算覆盖正确率
    coverage_rate = get / max_page if max_page > 0 else 0
    # accuracy_rate = get_high / max_page if max_page > 0 else 0
    # error_rate = (get - get_high) / max_page if max_page > 0 else 0

    print(f"覆盖正确率: {coverage_rate:.2%}")
    # print(f"正确率: {accuracy_rate:.2%}")
    # print(f"错误率: {error_rate:.2%}")

    # 输出一行四列的数据，用于复制到 Excel
    # print(f"{redundancy_rate:.4f}\t{coverage_rate:.4f}")

    # 按页码升序排列
    df = df.sort_values(by="页码", ascending=True)

    # 保存清洗后的数据到临时文件
    temp_file = input_file.replace(".xlsx", "_temp.xlsx")
    df.to_excel(temp_file, index=False, engine='openpyxl')

    # 使用 openpyxl 加载临时文件并进行高亮处理
    workbook = load_workbook(temp_file)
    sheet = workbook.active

    # 定义高亮样式
    highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")  # 黄色背景
    highlight_font = Font(color="FF0000")  # 红色字体
    low_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")  # 红色背景
    low_font = Font(color="FFFFFF")  # 白色字体

    # 寻找“精度”列
    columns = [sheet.cell(row=1, column=col).value for col in range(1, sheet.max_column + 1)]
    precision_col_index = None
    for col_name in columns:
        if "精度" in col_name:
            precision_col_index = columns.index(col_name) + 1
            break

    if precision_col_index is None:
        print("警告：未找到列名为'精度'的列，无法进行高亮处理。")
        os.remove(temp_file)  # 删除临时文件
        return

    # 遍历所有行进行高亮处理
    for row in range(2, sheet.max_row + 1):  # 从第2行开始，第一行是表头
        precision = sheet.cell(row=row, column=precision_col_index).value  # 获取“精度”列的值

        if precision == "中匹配度":
            # 高亮整行（黄色背景，红色字体）
            for col in range(1, len(columns) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.fill = highlight_fill
                cell.font = highlight_font
        elif precision == "低匹配度":
            # 高亮整行（红色背景，白色字体）
            for col in range(1, len(columns) + 1):
                cell = sheet.cell(row=row, column=col)
                cell.fill = low_fill
                cell.font = low_font

    # 保存到原始文件路径
    workbook.save(input_file)
    os.remove(temp_file)  # 删除临时文件
    print(f"处理完成，结果已保存到 {input_file}")


