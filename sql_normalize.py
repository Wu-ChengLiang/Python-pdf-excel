import pymssql
import pandas as pd
from openpyxl import load_workbook
import re
import copy
import json
from openpyxl import Workbook
from decimal import Decimal
import requests
from sqlalchemy import create_engine
import pandas as pd
import pymssql
import pdfplumber

# 定义查询 SQL 数据的函数
def querysql(EPBH):
    conn = pymssql.connect(host='10.106.22.60', user='GILDATA\wucl55091', password='hsjy@2025', database='JYCFI', charset='GBK')
    cursor = conn.cursor()
    sql = """
    SELECT
        A.EPBH AS [EP编号],
        A.HYDM AS [行业代码],
        A.XXLY AS [信息来源],
        A.XXLYBM [信息来源编码],
        LEFT(CONVERT(VARCHAR(10), A.XXFBRQ, 120), 10) AS [信息发布日期],
        LEFT(CONVERT(VARCHAR(10), A.JZRQ, 120), 10) AS [截止日期],
        A.CZND AS [财政年度],
        D.CHINAME AS [经营业务类型代码],
        E.CHINAME AS [数据类目一],
        A.SJLMYMC AS [数据类目一名称],
        A.SJLMYDM AS [数据类目一代码],
        F.CHINAME AS [数据类目二],
        A.SJLMEMC AS [数据类目二名称],
        A.SJLMEDM AS [数据类目二代码],
        G.CHINAME AS [数据类目三],
        A.SJLMSMC AS [数据类目三名称],
        A.ZTYSMC AS [主体原始名称],
        A.ZBDM AS [指标代码],
        Y.CHINAME AS [标准名称],
        A.ZBMC AS [指标名称],
        A.ZBSJ AS [指标数据],
        H.CHINAME AS [指标单位],
        A.BZSM AS [匹配代码-单位],
        J.CHINAME AS [统计口径],
        M.CHINAME AS [统计期间],
        A.YM AS [页码],
        A.BZSM AS [指标内容],
        A.SFYX AS [是否有效],
        A.BZSM AS [备注说明],
        A.HBM AS [行编码]
    FROM usrQYJYSJ A
    JOIN [10.106.22.51].JYPRIME.dbo.usrQYMB B ON B.EPBH = A.EPBH
    JOIN usrFCWXXMB L ON L.GILCODE = A.XXLYBM
    JOIN usrFCWXXMB D ON D.GILCODE = A.JYYWLXDM
    JOIN usrFCWXXMB J ON J.GILCODE = A.TJKJ
    JOIN usrFCWXXMB M ON M.GILCODE = A.TJQJ
    JOIN [10.106.22.51].JYPRIME.dbo.usrQYMB QM ON A.EPBH = QM.EPBH
    LEFT JOIN usrFCWXXMB E ON E.GILCODE = A.SJLMY
    LEFT JOIN usrFCWXXMB F ON F.GILCODE = A.SJLME
    LEFT JOIN usrFCWXXMB G ON G.GILCODE = A.SJLMS
    LEFT JOIN usrFCWXXMB H ON H.GILCODE = A.ZBDW
    LEFT JOIN usrFCWXXMB I ON I.GILCODE = A.BZ
    LEFT JOIN usrFCWXXMB K ON K.GILCODE = A.SFYX
    LEFT JOIN usrFCWXXMB X ON X.GILCODE = A.HYDM
    LEFT JOIN usrFCWXXMB Y ON Y.GILCODE = A.ZBDM
    WHERE A.EPBH IN (%s)
    ORDER BY A.XGSJ DESC
    """
    cursor.execute(sql, (EPBH,))
    data = cursor.fetchall()
    cursor.close()
    conn.close()
    return data


#sql标准化（只是标准化括号，以及统一之后不会改变数据特性的部分）
# 加载映射表
#因为数据有: 一定要替换中文符号
def load_mapping_table_sql(file_path):
    workbook = load_workbook(filename=file_path)
    sheet = workbook.active
    mapping_dict = {}
    # 遍历工作表的每一行，从第2行开始（跳过表头）
    for row in sheet.iter_rows(min_row=2, values_only=True):
        # 将所有数据转换为字符串，并清理空格
        original_name = str(row[0]).strip() if row[0] is not None else None
        standardized_name = str(row[1]).strip() if row[1] is not None else None

        # 替换特殊字符（如全角冒号）
        original_name = original_name.replace('：', ':') if original_name else None
        standardized_name = standardized_name.replace('：', ':') if standardized_name else None

        # 检查是否为有效数据
        if original_name and standardized_name:
            mapping_dict[original_name] = standardized_name
        else:
            # 如果某行无效，跳过该行，继续读取下一行
            continue
    return mapping_dict


# sql数据清洗和标准化
def clean_and_standardize_data(data,fields_to_clean, mapping_dict):
    # 标准格式，不需要颜色注记和附加的三列，但是保留报错提示
    columns = [
        "EP编号", "行业代码", "信息来源", "信息来源编码", "信息发布日期", "截止日期", "财政年度",
        "经营业务类型代码", "数据类目一", "数据类目一名称", "数据类目一代码", "数据类目二",
        "数据类目二名称", "数据类目二代码", "数据类目三", "数据类目三名称", "主体原始名称",
        "指标代码", "标准名称", "指标名称", "指标数据", "指标单位", "匹配代码-单位",
        "统计口径", "统计期间", "页码", "指标内容", "是否有效", "备注说明", "行编码"
    ]

    cleaned_data = []
    for row in data:
        cleaned_row = {}
        for i, key in enumerate(columns):
            item = row[i]
            if key in fields_to_clean and isinstance(item, str):
                # 数据清洗：简单替换
                cleaned_item = item.replace('（', '(').replace('）', ')').replace(' ','').replace('  ','').replace('：',':')
            else:
                cleaned_item = item
            cleaned_row[key] = cleaned_item
        cleaned_data.append(cleaned_row)
    return cleaned_data







