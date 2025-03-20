import pdfplumber
from openpyxl import load_workbook
from decimal import Decimal

def load_mapping_table(pdf_path):
    # 加载映射表
        workbook = load_workbook(filename=pdf_path)
        sheet = workbook.active
        mapping_dict = {}
        # 遍历工作表的每一行，从第2行开始（跳过表头）
        for row in sheet.iter_rows(min_row=2, values_only=True):
            original_name = row[2]  # 第2列：pdf原始名称
            standardized_name = row[3]  # 第3列：pdf标准化名称
            # 检查是否为有效数据
            if original_name is not None and standardized_name is not None:
                mapping_dict[original_name] = standardized_name
            else:
                # 如果某行无效，停止读取后续行
                break
        return mapping_dict

#提取 PDF 文件中的表格数据并进行清洗和标准化。
def extract_clean_pdf(pdf_file, mapping_dict,start_page=10, end_page=None):
    cleaned_data = []
    with pdfplumber.open(pdf_file) as pdf:  # 打开 PDF 文件
        #如果未指定结束页，默认处理所有页面
        if end_page is None:
            end_page = len(pdf.pages)

        for page in pdf.pages[start_page:end_page]:# 遍历 PDF 的指定页面

            tables = page.extract_tables()  # 提取该页的全部表格
            for table in tables:  # 遍历每一页中的表格
                for row in table:  # 遍历表格的每一行
                    # 清洗每一行的数据：替换换行符、逗号、中文括号，并去除空值和仅包含空格的单元格
                    cleaned_row = [
                        cell.replace('\n', '').replace(',', '').replace('（', '(').replace('）', ')').replace(' ', '')
                        if cell is not None else ''
                        for cell in row
                    ]

                    #cell.strip 移除空白字符  直接移除可能导致错位，这个要用后续的手段弥补  #数据的提取和后续的处理是一体化的过程
                    cleaned_row = [cell for cell in cleaned_row if cell.strip()]  # 删除空值和仅包含空格的单元格
                    # 标准化数据：根据映射表替换

                    standardized_row = []
                    for cell in cleaned_row:
                        if cell in mapping_dict:
                            standardized_row.append(mapping_dict[cell])
                        else:
                            standardized_row.append(cell)
                    # 将标准化后的行添加到结果中
                    cleaned_data.append(standardized_row)
    return cleaned_data

#————————————————————————————————————上海 2.经营表——————————————————————————————————————————————————————
def extract_pdf_table2_shanghai(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    extracted_data = []
    """
    从PDF数据中提取表格数据，支持强匹配、弱匹配和最终备用方案。
    参数说明：
    - pdf_data: 二维列表，从PDF提取的数据。
    - start_pattern1: 第一组强匹配模式。
    - start_pattern2: 第二组弱匹配模式。
    - start_pattern3: 第三组最终备用模式。
    - end_pattern: 结束模式的正则表达式。
    - 强匹配最大行数: 强匹配时的最大行数限制。
    - 弱匹配最大行数: 弱匹配时的最大行数限制。
    - 弱匹配偏移量: 弱匹配时的起始偏移量。
    - 备用偏移量: 最终备用方案的起始偏移量。
    """
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"经营表格没有分产品,只能采集分行业")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"尴尬，找不到任何匹配的2.经营表")

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围，索引+1，即代表索引开始的下一个开始截取；+0即代表从此位置开始截取
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 1:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index + 1:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]

    return extracted_data

def analysis_pdf_table2_shanghai(pattern):
    pdf_dic = []  # 初始化结果列表
    heji_count = 0  # 初始化合计行计数器
    for row in pattern:  # 遍历子数组
        # 确保每行有7个元素，不足部分用'0'补全
        while len(row) < 7:
            row.append('0')
        # 检查是否是合计行，并进行特殊处理
        if row[0] == "合计" or row[0] == "小计" or row[0] == "主营业务合计" :  # 检查第0列是否为“合计”
            heji_count += 1  # 合计行计数器加1
            if heji_count == 1:
                row[0] = "产品合计"  # 第一次碰到合计，处理为“产品合计”
            elif heji_count == 2:
                row[0] = "地区合计"  # 第二次碰到合计，处理为“地区合计”
            elif heji_count == 3:
                row[0] = "渠道合计"  # 第三次碰到合计，处理为“渠道合计”
            elif heji_count == 4:
                row[0] = "合计"
            else:
                row[0] = "?其他合计?"  # 如果还有其他合计行，统一处理为“其他合计”,可能遇到了特殊情况
        # 提取数据
        income_increase = row[4].replace('%', '')
        expense_increase = row[5].replace('%', '')
        profit_margin = row[6].replace('增加', '').replace('上升', '').replace('减少', '-').replace('下降', '-').replace('个百分点', '').replace('个', '').replace('百分点', '')
        data = {
            "数据类目名称": row[0],
            "营业收入": row[1],
            "营业成本": row[2],
            "毛利率": row[3],
            "营业收入比上年增减": income_increase,
            "营业成本比上年增减": expense_increase,
            "毛利率比上年增减": profit_margin
        }
        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表

#——————————————————————————————————————————————上海 3.产销表——————————————————————————————————————————————————————————
def extract_pdf_table3_shanghai(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    extracted_data = []
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"尴尬，找不到任何匹配的3.产销表")

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 1:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index + 1:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]

    return extracted_data


def analysis_pdf_table3_shanghai(pattern):
    pdf_dic = []  # 初始化结果列表
    for row in pattern:  # 遍历子数组
        #产销表的合集行，不必特殊处理
        # 确保每行有8个元素，不足部分用'0'补全
        while len(row) < 8:
            row.append('0')
        # 提取数据
        income_increase = row[5].replace('%', '')
        expense_increase = row[6].replace('%', '')
        profit_margin = row[7].replace('增加', '')#要不要把/替换成None？？
        data = {
            "主要产品": row[0],
            "单位": row[1],
            "生产量": row[2],
            "销售量": row[3],
            "库存量": row[4],
            "生产量比上年增减": income_increase,
            "销售量比上年增减": expense_increase,
            "库存量比上年增减": profit_margin
        }

        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表



#——————————————————————————————————————————————上海 4.成本分析表——————————————————————————————————————————————————————————

def extract_pdf_table4_shanghai(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    # 初始化变量
    extracted_data = []
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"尴尬，找不到任何匹配的4.成本分析表")

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 1:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index + 1:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]

    if end_index == None:
        end_index = 12

    return extracted_data,end_index


# def analysis_pdf_table4_shanghai(pattern):
#     pdf_dic = []  # 初始化结果列表
#     last_valid_first_element = None  # 用于记录最近一次有效的第0列值
#     # 下面是原来的数据程序，但是因为上面修改了，逻辑要改成：
#     # 如果[1]是空的，填入最近一次有效的 [0] 和 [1] 都有值的 [0]
#
#
#     def is_number(value):
#         try:
#             float(value)  # 尝试将值转换为浮点数
#             return True
#         except ValueError:
#             return False
#
#     # 遍历并修改数据
#     for row in pattern:
#         if ('分产品' in row and '成本构成项目' in row) or ('分行业' in row and '成本构成项目' in row):
#             continue  # 跳过特定行
#
#         # 检查最后一列是否为数字
#
#         while len(row) > 0 and not is_number(row[-1]):  # 如果不是数字，循环移除最后一列
#             row.pop()  # 使用 pop() 方法移除最后一列
#
#     # 获取最大列数目
#     max_columns = 0
#     for row in pattern:
#             max_columns = max(max_columns, len(row))
#     for row in pattern:  # 遍历子数组
#
#         if len(row) == max_columns:
#             last_valid_first_element = row[0]  # 更新最近一次有效的第0列值
#         elif len(row) == max_columns - 1:
#             if last_valid_first_element is not None:
#                 row.insert(0, last_valid_first_element)  # 在第0列插入值
#         else:
#             # 如果行长度小于 max_columns - 1，在前面补默认值
#             row = ['默认值'] * (max_columns - len(row)) + row
#             # 确保 row 的长度达到 max_columns
#         while len(row) < max_columns:
#             row.append('默认值')
#
#         # 提取数据
#         data = {
#             "分产品": row[0],
#             "成本构成项目": row[1],
#             "本期金额": row[2],
#             "本期占总成本比例": row[3],
#             "本期金额较上年同期变动比例": row[6]
#         }
#         pdf_dic.append(data)  # 将字典添加到结果列表
#     return pdf_dic  # 返回结果列表

def analysis_pdf_table4_shanghai(pattern):
    pdf_dic = []  # 初始化结果列表
    last_valid_first_element = None  # 用于记录最近一次有效的第0列值

    def is_number(value):
        try:
            float(value)  # 尝试将值转换为浮点数
            return True
        except ValueError:
            return False

    # 遍历并修改数据
    for row in pattern:
        if ('分产品' in row and '成本构成项目' in row) or ('分行业' in row and '成本构成项目' in row):
            continue  # 跳过特定行

        # 检查最后一列是否为数字
        while len(row) > 0 and not is_number(row[-1]):  # 如果不是数字，循环移除最后一列
            row.pop()  # 使用 pop() 方法移除最后一列

    # 获取最大列数目
    max_columns = 0
    for row in pattern:
        max_columns = max(max_columns, len(row))

    for row in pattern:  # 遍历子数组
        current_length = len(row)
        if current_length == 0:
            continue  # 跳过空行

        # 动态调整列数，确保每行都达到 max_columns 列
        while len(row) < max_columns:
            row.append('默认值')  # 填充默认值

        # 提取数据时增加对列数的检查
        if len(row) >= 7:  # 确保至少有7列数据
            data = {
                "分产品": row[0] if len(row) > 0 else '默认值',
                "成本构成项目": row[1] if len(row) > 1 else '默认值',
                "本期金额": row[2] if len(row) > 2 else '默认值',
                "本期占总成本比例": row[3] if len(row) > 3 else '默认值',
                "本期金额较上年同期变动比例": row[6] if len(row) > 6 else '默认值'
            }
        else:
            # 如果列数不足7，填充默认值
            data = {
                "分产品": row[0] if len(row) > 0 else '默认值',
                "成本构成项目": row[1] if len(row) > 1 else '默认值',
                "本期金额": row[2] if len(row) > 2 else '默认值',
                "本期占总成本比例": row[3] if len(row) > 3 else '默认值',
                "本期金额较上年同期变动比例": '默认值'
            }

        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表


#——————————————————————————————————————————————上海 5.前五大客户表——————————————————————————————————————————————————————————
def analysis_pdf_table5_shanghai(pattern):
    pdf_dic = []  # 初始化结果列表
    # 检查 pattern 是否为空
    if not pattern:
        return []
    for row in pattern:  # 遍历子数组
        # 确保每行有6个元素，不足部分用'?缺失'补全
        while len(row) < 6:
            row.append('?缺失')
        if row[0] == "合计":  # 检查第0列是否为“合计”
            row[1] = "前五大客户合计"  # 将第1列设置为“合计”
        data = {
            "客户名称": row[1],
            "销售额": row[2],
            "占年度销售总额比例": row[3]
        }
        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表


def extract_pdf_table5_shanghai(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    extracted_data = []
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"未披露5.前五大客户表,尝试文本方式")
            return "5.text_start_signal"


    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]
    return extracted_data


#——————————————————————————————————————————————上海 6.前五大供应商表——————————————————————————————————————————————————————————
def analysis_pdf_table6_shanghai(pattern):
    pdf_dic = []  # 初始化结果列表
    # 检查 pattern 是否为空
    if not pattern:
        return []
    for row in pattern:  # 遍历子数组
        # 确保每行有6个元素，不足部分用'?缺失'补全
        while len(row) < 6:
            row.append('?缺失')
        if row[0] == "合计":  # 检查第0列是否为“合计”
            row[1] = "前五大供应商合计"  # 将第1列设置为“合计”
        data = {
            "供应商名称": row[1],
            "采购额": row[2],
            "占年度采购总额比例": row[3]
        }
        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表


def extract_pdf_table6_shanghai(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    extracted_data = []
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"未披露6.前五大供应商表，尝试文本方式")
            return "6.text_start_signal"

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]
    return extracted_data


#——————————————————————————————————————上海 7.专利表  7.报告期内获得的知识产权列表———————————————————————————————————————————————————
def analysis_pdf_table7_shanghai(pattern):
    pdf_dic = []  # 初始化结果列表
    # 检查 pattern 是否为空
    if not pattern:
        return []
    for row in pattern:  # 遍历子数组
        # 确保每行有6个元素，不足部分用'?缺失'补全
        while len(row) < 6:
            row.append('?缺失')

        if row[0] == "合计":  # 检查第0列是否为“合计”
            row[0] = "默认值"  # 将第1列设置为“”,！！！！！！！！！！！！！！！！！测试异常看这里
        data = {
            "专利类型": row[0],
            "新增知识产权申请数量": row[1],
            "新增知识产权有效数量": row[2],
            "累计知识产权申请数量": row[3],
            "累计知识产权有效数量": row[4]

        }
        pdf_dic.append(data)  # 将字典添加到结果列表
    return pdf_dic  # 返回结果列表


def extract_pdf_table7_shanghai(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    extracted_data = []
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"尴尬，找不到任何匹配的7.专利表")

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]

    return extracted_data


# ————————————————————————————————————深圳 2.经营表  经营2表协同匹配，前10%表从经营表之后的位置开始匹配——————————————————————————————————————————————————————
def extract_pdf_table2_shenzhen(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    extracted_data = []
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"经营表格没有分产品,只能采集分行业")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"尴尬，找不到任何匹配的2.经营表")

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    #协同匹配:传递end_index 如果未找到结束模式，则设置 end_index 为 first_index + 8
    if end_index is None:
        if first_index is not None:
            end_index = first_index + 8
        else:
            end_index = None


    # 根据匹配类型截取字段范围，索引+1，即代表索引开始的下一个开始截取；+0即代表从此位置开始截取
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 1:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index + 1:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]

    return extracted_data,end_index


def analysis_pdf_table2_shenzhen(pattern):
    pdf_dic = []  # 初始化结果列表
    heji_count = 0  # 初始化合计行计数器
    for row in pattern:  # 遍历子数组
        # 确保每行有7个元素，不足部分用'0'补全
        while len(row) < 6:
            row.append('0')

        # 检查是否是合计行，并进行特殊处理
        if row[0] == "合计" or row[0] == "小计" or row[0] == "主营业务合计":  # 检查第0列是否为“合计”
            heji_count += 1  # 合计行计数器加1
            if heji_count == 1:
                row[0] = "产品合计"  # 第一次碰到合计，处理为“产品合计”
            elif heji_count == 2:
                row[0] = "地区合计"  # 第二次碰到合计，处理为“地区合计”
            elif heji_count == 3:
                row[0] = "渠道合计"  # 第三次碰到合计，处理为“渠道合计”
            elif heji_count == 4:
                row[0] = "合计"
            else:
                row[0] = "?其他合计?"  # 如果还有其他合计行，统一处理为“其他合计”,可能遇到了特殊情况

        # 提取数据
        income_increase = row[5].replace('%', '')
        data = {
            "数据类目名称": row[0].replace('%',''),
            "营业收入": row[1].replace('%',''),
            "营业收入比上年增减": income_increase,
        }
        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表


# ————————————————————————————————————深圳 2.5 前10%经营表——————————————————————————————————————————————————————
def extract_pdf_table2_5_shenzhen(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量,start_index=None):
    extracted_data = []
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 协同匹配，如果提供了起始索引，则从起始索引开始匹配；否则从头开始
    if start_index is not None:
        range_pattern = pdf_data[start_index + 1:]
    else:
        range_pattern = pdf_data

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(range_pattern):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(range_pattern):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"经营表格没有分产品,只能采集分行业")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(range_pattern):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"尴尬，找不到任何匹配的2.5 前10%经营表")

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(range_pattern)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in range_pattern[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围，索引+1，即代表索引开始的下一个开始截取；+0即代表从此位置开始截取
    if match_type == "strong":
        if end_index is not None:
            extracted_data = range_pattern[first_index + 1:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = range_pattern[first_index + 1:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = range_pattern[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = range_pattern[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = range_pattern[first_index + 备用偏移量:end_index]
        else:
            extracted_data = range_pattern[first_index + 备用偏移量:first_index + 弱匹配最大行数]

    return extracted_data


def analysis_pdf_table2_5_shenzhen(pattern):
    pdf_dic = []  # 初始化结果列表
    heji_count = 0  # 初始化合计行计数器
    for row in pattern:  # 遍历子数组
        # 确保每行有7个元素，不足部分用'0'补全
        while len(row) < 7:
            row.append('0')
        # 检查是否是合计行，并进行特殊处理
        if row[0] == "合计" or row[0] == "小计" or row[0] == "主营业务合计":  # 检查第0列是否为“合计”
            heji_count += 1  # 合计行计数器加1
            if heji_count == 1:
                row[0] = "产品合计"  # 第一次碰到合计，处理为“产品合计”
            elif heji_count == 2:
                row[0] = "地区合计"  # 第二次碰到合计，处理为“地区合计”
            elif heji_count == 3:
                row[0] = "渠道合计"  # 第三次碰到合计，处理为“渠道合计”
            elif heji_count == 4:
                row[0] = "合计"
            else:
                row[0] = "?其他合计?"  # 如果还有其他合计行，统一处理为“其他合计”,可能遇到了特殊情况

        # 提取数据
        income_increase = row[4].replace('%', '')
        expense_increase = row[5].replace('%', '')
        profit_margin = row[6].replace('%', '').replace('增加', '').replace('上升', '').replace('减少', '-').replace('下降', '-').replace('个百分点', '').replace('个', '').replace('百分点', '')
        data = {
            "数据类目名称": row[0].replace('%', ''),
            "营业收入": row[1].replace('%', ''),
            "营业成本": row[2].replace('%', ''),
            "毛利率": row[3].replace('%', ''),
            "营业收入比上年增减": income_increase,
            "营业成本比上年增减": expense_increase,
            "毛利率比上年增减": profit_margin
        }
        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表



# ——————————————————————————————————————————————深圳 3.产销表——————————————————————————————————————————————————————————
def extract_pdf_table3_shenzhen(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    extracted_data = []
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"尴尬，找不到任何匹配的3.产销表")

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 1:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index + 1:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]

    return extracted_data


def analysis_pdf_table3_shenzhen(pattern):
    pdf_dic = []  # 初始化结果列表
    last_valid_first_element = None  # 用于记录最近一次有效的第0列值
    for row in pattern:  # 遍历子数组
        # 如果当前行有6个元素，更新最近一次有效的第0列值
        if len(row) == 6:
            last_valid_first_element = row[0]  # 更新最近一次有效的第0列值
        elif len(row) == 5:
            # 如果当前行有5个元素，从最近一次有效的第0列值中取值
            if last_valid_first_element is not None:
                row.insert(0, last_valid_first_element)  # 在第0列插入值
            else:
                # 如果没有找到有效的第0列值，用0填充
                row.insert(0, '0')
        else:
            # 如果当前行元素少于5个，用0补全到6个元素
            if len(row) < 5:
                row = ['0'] * (6 - len(row)) + row  # 在前面补0
        # 确保每行现在都有5个元素
        while len(row) < 5:
            row.append('0')
        # 提取数据
        prodsale_proportion = row[3]
        prodsale_change = row[5].replace('%', '') if '%' in row[5] else row[5]
        data = {
            "行业分类": row[0],
            "项目": row[1],
            "单位": row[2],
            "生产量、销售量、库存量": prodsale_proportion,
            "生产量、销售量、库存量比上年增减": prodsale_change
        }
        pdf_dic.append(data)  # 将字典添加到结果列表
    return pdf_dic  # 返回结果列表


#3.产销表

def analysis_pdf_table3_shenzhen(pattern):
    pdf_dic = []  # 初始化结果列表
    last_valid_first_element = None  # 用于记录最近一次有效的第0列值

    def is_number(value):
        try:
            float(value)  # 尝试将值转换为浮点数
            return True
        except ValueError:
            return False

    # 遍历并修改数据
    for row in pattern:
        if ('行业分类' in row and '单位' in row) or ('产品分类' in row and '单位' in row):
            continue  # 跳过特定行

        #深交所的最后一列并不是备用字段，所以当我把这一列移除的时候，把数据也也移除了
        # # 检查最后一列是否为数字
        # while len(row) > 0 and not is_number(row[-1]):  # 如果不是数字，循环移除最后一列
        #     row.pop()  # 使用 pop() 方法移除最后一列

    # 获取最大列数目
    max_columns = 0
    for row in pattern:
        max_columns = max(max_columns, len(row))
    for row in pattern:  # 遍历子数组
        if len(row) == 6:
            last_valid_first_element = row[0]  # 更新最近一次有效的第0列值
        elif len(row) == 5:
            if last_valid_first_element is not None:
                row.insert(0, last_valid_first_element)  # 在第0列插入值
        else:
            # # 如果行长度小于 max_columns - 1，在前面补默认值
            # row = row  + ['默认值'] * (max_columns - len(row))
            # # 确保 row 的长度达到 max_columns
            pass
        while len(row) < 6:
            row.append('默认值')

        # 深交所的产销表结构不一致
        # 提取数据
        prodsale_change = row[5].replace('%', '')
        data = {
            "行业分类": row[0],
            "项目": row[1],
            "单位": row[2],
            "生产量、销售量、库存量": row[3],
            "生产量、销售量、库存量比上年增减": prodsale_change,
        }
        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表



# ——————————————————————————————————————————————深圳 4.成本分析表——————————————————————————————————————————————————————————
def extract_pdf_table4_shenzhen(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    # 初始化变量
    extracted_data = []
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"尴尬，找不到任何匹配的4.成本分析表")

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 1:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index + 1:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]

    if end_index == None:
        end_index = 12

    return extracted_data,end_index

def analysis_pdf_table4_shenzhen(pattern):
    pdf_dic = []  # 初始化结果列表
    last_valid_first_element = None  # 用于记录最近一次有效的第0列值

    import re
    def is_str(value):
        repattern=re.compile(r"\D.*")
        # 使用正则表达式判断是否为纯文字
        if isinstance(value,str) and repattern.match(value):
            return True
        return False

    # 遍历并修改数据
    for row in pattern:
        if ('产品分类' in row and '项目' in row) or ('行业分类' in row and '项目' in row):
            continue  # 跳过特定行

        # 检查前两列是否为文本
        if len(row) >1:
            if not is_str(row[0]) or not is_str(row[1]):
                row.insert(1, '')

    # 获取最大列数目
    max_columns = 0
    for row in pattern:
        max_columns = max(max_columns, len(row))




    for row in pattern:  # 遍历子数组
        # 下面是原来的数据程序，但是因为上面修改了，逻辑要改成：
        # 如果[1]是空的，填入最近一次有效的 [0] 和 [1] 都有值的 [0]
        while len(row) < 7:
                row.append('')
        if row[1] :    #如果 [0] 和 [1]都有值，储存数值
            last_valid = row[0]   #填入这个值
        elif row[1] == "" :
            row.insert(0, last_valid)
        # 提取数据
        data = {
            "分产品": row[0],
            "成本构成项目": row[1],
            "本期金额": row[2],
            "本期占总成本比例": row[3],
            "本期金额较上年同期变动比例": row[6]
        }
        pdf_dic.append(data)  # 将字典添加到结果列表
    return pdf_dic  # 返回结果列表

# ——————————————————————————————————————————————深圳 5.前五大客户表——————————————————————————————————————————————————————————
def analysis_pdf_table5_shenzhen(pattern):
    pdf_dic = []  # 初始化结果列表
    for row in pattern:  # 遍历子数组
        # 确保每行有6个元素，不足部分用'?缺失'补全
        while len(row) < 6:
            row.append('?缺失')
        # 确保每行有6个元素，不足部分用'?缺失'补全
        if row[0] == "合计":  # 检查第0列是否为“合计”
            row[1] = "前五大客户合计"  # 将第1列设置为“合计”

        data = {
            "客户名称": row[1],
            "销售额": row[2],
            "占年度销售总额比例": row[3]
        }
        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表


def extract_pdf_table5_shenzhen(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    extracted_data = []
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"尴尬，找不到任何匹配的6.前五大客户表")

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]
    return extracted_data


# ——————————————————————————————————————————————深圳 6.前五大供应商表——————————————————————————————————————————————————————————
def analysis_pdf_table6_shenzhen(pattern):
    pdf_dic = []  # 初始化结果列表
    for row in pattern:  # 遍历子数组
        # 确保每行有6个元素，不足部分用'?缺失'补全
        while len(row) < 6:
            row.append('?缺失')

        if row[0] == "合计":  # 检查第0列是否为“合计”
            row[1] = "前五大供应商合计"  # 将第1列设置为“合计”

        data = {
            "供应商名称": row[1],
            "采购额": row[2],
            "占年度采购总额比例": row[3]
        }
        pdf_dic.append(data)  # 将字典添加到结果列表

    return pdf_dic  # 返回结果列表


def extract_pdf_table6_shenzhen(pdf_data, start_pattern1, start_pattern2, start_pattern3, end_pattern,
                                强匹配最大行数, 弱匹配最大行数, 弱匹配偏移量, 备用偏移量):
    extracted_data = []
    # 初始化变量
    first_index = None  # 用于记录第一个模式的索引
    end_index = None  # 用于记录结束模式的索引
    match_type = None  # 用于记录匹配类型（强匹配、弱匹配或最终备用）

    # 尝试第一组模式（强匹配）
    for i, row in enumerate(pdf_data):
        if len(row) >= len(start_pattern1) and all(pattern.match(cell) for pattern, cell in zip(start_pattern1, row)):
            first_index = i
            match_type = "strong"
            break

    # 如果第一组模式没有匹配成功，尝试第二组模式（弱匹配）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern2) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern2, row)):
                first_index = i
                match_type = "weak"
                break
        else:
            print(f"")

    # 如果前两组模式都没有匹配成功，尝试第三组模式（最终备用方案）
    if first_index is None:
        for i, row in enumerate(pdf_data):
            if len(row) >= len(start_pattern3) and all(
                    pattern.match(cell) for pattern, cell in zip(start_pattern3, row)):
                first_index = i
                match_type = "final_fallback"
                break
        else:
            print(f"6.前五大供应商表，尝试文本方式")
            return "6.text_start_signal"

    # 如果找到了第一个模式，继续查找结束模式
    if first_index is not None:
        for k in range(first_index + 1, len(pdf_data)):  # 从第一个模式的下一行开始查找
            if any(end_pattern.search(cell) for cell in pdf_data[k]):  # 匹配整行
                end_index = k
                break

    # 根据匹配类型截取字段范围
    if match_type == "strong":
        if end_index is not None:
            extracted_data = pdf_data[first_index:end_index]  # 从第一个模式到结束模式
        else:
            extracted_data = pdf_data[first_index:first_index + 强匹配最大行数]

    elif match_type == "weak":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 弱匹配偏移量:first_index + 弱匹配最大行数]

    elif match_type == "final_fallback":
        if end_index is not None:
            extracted_data = pdf_data[first_index + 备用偏移量:end_index]
        else:
            extracted_data = pdf_data[first_index + 备用偏移量:first_index + 弱匹配最大行数]
    return extracted_data


#————————————————————————————————————————文字版本抽取方案 5.————————————————
def text_analysis_pdf_table5(pdf_path, start_page=None, end_page=None):
    """
    从 PDF 文件中提取并分类采购信息
    """
    # 定义匹配模式
    patterns = [  # 第一个为匹配字符，第二个为标准名称，第三个为数据类目一名称
        # (r'前五名供应商采购额(\d+.\d{2})万元，占年度采购总额(\d+.\d{2})%', '占年度采购总额比例', '采购金额占比'),
        (r'前五名客户销售额(\d+.\d{2})万元，占年度销售总额(\d+.\d{2})%', '占年度销售总额比例', '销售金额占比'),
    ]

    result = []

    # 读取 PDF 文件
    import pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        if end_page is None:
            end_page = len(pdf.pages)

        # 遍历 PDF 的指定页面
        for page_num in range(start_page, end_page):
            page = pdf.pages[page_num]

            # 提取文本
            text = page.extract_text()
            if text:
                cleaned_text = clean_text(text)
                import re
                # 遍历每个模式
                for pattern, standard_name, category in patterns:
                    matches = re.findall(pattern, cleaned_text)
                    if matches:                          #[('570363.90', '33.17')]
                        purchase_amount = matches[0][0]  #.replace(',', '')  # 去除千位分隔符
                        percentage = matches[0][1]
                        result.append({
                                '客户名称': '前五大客户合计',
                                '销售额': purchase_amount,
                                '占年度销售总额比例': f'{percentage}%'
                            })

    return result
#——————————————————————————————————————深圳文本抽取方案:6.————————————————————————————————
def text_analysis_pdf_table6(pdf_path, start_page=None, end_page=None):
    """
    从 PDF 文件中提取并分类采购信息
    """
    # 定义匹配模式
    patterns = [  # 第一个为匹配字符，第二个为标准名称，第三个为数据类目一名称
        # (r'前五名供应商采购额(\d+.\d{2})万元', '采购额', '采购金额'),
        (r'前五名供应商采购额(\d+.\d{2})万元，占年度采购总额(\d+.\d{2})%', '占年度采购总额比例', '采购金额占比'),
    ]

    result = []

    # 读取 PDF 文件
    import pdfplumber
    with pdfplumber.open(pdf_path) as pdf:
        if end_page is None:
            end_page = len(pdf.pages)

        # 遍历 PDF 的指定页面
        for page_num in range(start_page, end_page):
            page = pdf.pages[page_num]

            # 提取文本
            text = page.extract_text()
            if text:
                cleaned_text = clean_text(text)
                import re
                # 遍历每个模式
                for pattern, standard_name, category in patterns:
                    matches = re.findall(pattern, cleaned_text)
                    if matches:                          #[('570363.90', '33.17')]
                        purchase_amount = matches[0][0]  #.replace(',', '')  # 去除千位分隔符
                        percentage = matches[0][1]
                        result.append({
                                '供应商名称': '前五大供应商合计',
                                '采购额': purchase_amount,
                                '占年度采购总额比例': f'{percentage}%'
                            })

    return result

# pdf_path = r'C:\Users\wucl55091\Desktop\Juno-win32.win32.x86_64\files\765053993908.PDF'

#——————————————————————————————————————深圳 7.专利表  7.报告期内获得的知识产权列表———————————————————————————————————————————————————
#文本形式

def clean_text(text):
    """
    清洗文本：替换换行符、逗号、中文括号，但保留空格
    """
    if text is None:
        return ''
    text = text.replace('\n', '').replace(',', '').replace('（', '(').replace('）', ')').replace(' ','')
    return text.strip()

def analysis_pdf_table7_shenzhen(pdf_file, start_page=None, end_page=None):
    """
    从 PDF 文件中提取并分类专利信息
    """
    # 1.在每一条pattern+时间前缀  目前方案1
    # 2.遍历多个pattern    #可能速度更慢

    pattern1 = [
        (r'累计获得授权专利(\d+)件', '累计知识产权有效数量', '专利'),
        (r'累计获得授权专利.*?发明专利\s*(\d+)\s*件', '累计知识产权有效数量', '发明专利'),
        (r'累计获得授权专利.*?实用新型\s*(\d+)\s*件', '累计知识产权有效数量', '实用新型专利'),
    ]
    pattern2 = [
        (r'正在申请的专利为(\d+)件', '新增知识产权申请数量', '专利'),
        (r'正在申请的专利为.*?发明专利\s*(\d+)\s*件', '新增知识产权申请数量', '发明专利'),
        (r'正在申请的专利为.*?实用新型\s*(\d+)\s*件', '新增知识产权申请数量', '实用新型专利'),
    ]
    patterns = pattern1 + pattern2

    # 初始化数据结构
    patent_data = {
        '发明专利': {'专利类型': '发明专利', '新增知识产权申请数量': '', '新增知识产权有效数量': '', '累计知识产权申请数量': '', '累计知识产权有效数量': ''},
        '实用新型专利': {'专利类型': '实用新型专利', '新增知识产权申请数量': '', '新增知识产权有效数量': '', '累计知识产权申请数量': '', '累计知识产权有效数量': ''},
        '外观设计专利': {'专利类型': '外观设计专利', '新增知识产权申请数量': '', '新增知识产权有效数量': '', '累计知识产权申请数量': '', '累计知识产权有效数量': ''},
        '软件著作权': {'专利类型': '软件著作权', '新增知识产权申请数量': '', '新增知识产权有效数量': '', '累计知识产权申请数量': '', '累计知识产权有效数量': ''},
        '其他': {'专利类型': '其他', '新增知识产权申请数量': '', '新增知识产权有效数量': '', '累计知识产权申请数量': '', '累计知识产权有效数量': ''},
        '专利': {'专利类型': '专利', '新增知识产权申请数量': '', '新增知识产权有效数量': '', '累计知识产权申请数量': '', '累计知识产权有效数量': ''}
    }

    # 读取 PDF 文件
    with pdfplumber.open(pdf_file) as pdf:
        if end_page is None:
            end_page = len(pdf.pages)

        # 遍历 PDF 的指定页面
        for page_num in range(start_page, end_page):
            page = pdf.pages[page_num]

            # 提取文本
            text = page.extract_text()
            if text:
                cleaned_text = clean_text(text)
                import re
                # 匹配文本中的专利信息
                for pattern, standard_name, category in patterns:
                    matches = re.findall(pattern,cleaned_text)
                    for match in matches:
                        if match.isdigit():
                            current_value = patent_data[category][standard_name]
                            if not current_value or int(match) > int(current_value):
                                patent_data[category][standard_name] = match

    # 将数据整理成最终的列表形式
    result = []
    for category in patent_data:
        result.append(patent_data[category])

    return result






